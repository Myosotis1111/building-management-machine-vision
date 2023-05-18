"""
-------------------------------------------------
Project Name: Building Management: Machine Vision
File Name: detection_platform.py
Author: Xinchen Yang
Last Modified Date: 05/05/2023
Description：used for roller shutter detection in TH Luebeck.
-------------------------------------------------
"""

import os
import io
import shutil
from datetime import datetime, timedelta
from pathlib import Path
import cv2
import openpyxl
import numpy as np
import torch
import matplotlib.pyplot as plt
from PySide2.QtCore import QDateTime, QSize
from PySide2.QtGui import QImage, QPixmap, Qt, QMouseEvent, QIcon
from PySide2.QtUiTools import QUiLoader
from PySide2.QtWidgets import QApplication, QFileDialog, QMessageBox, QDialog, QLabel, QFrame
import os.path as osp
from PIL import Image
import functools
from models.common import DetectMultiBackend
from utils.datasets import LoadImages
from utils.general import check_img_size, scale_coords, LOGGER, non_max_suppression
from utils.plots import Annotator, colors
from utils.torch_utils import select_device, time_sync
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Font, Side, Alignment
from tkinter import filedialog, messagebox
import tkinter as tk
import string


class Window:
    def __init__(self, doc, status, doc_max=None, doc_min=None, count_open=0, count_closed=0):
        self._doc = doc  # The degree of Closure of the roller shutter
        self._status = status  # The status of the roller shutter, can be open, closed, blocked or not_detected
        self._doc_max = doc_max  # The maximum possible DoC
        self._doc_min = doc_min  # The minimum possible DoC
        self.button = None  # Corresponding button No. in GUI.
        self._count_open = count_open  # The number of open status detected through the detection period
        self._count_closed = count_closed  # The number of close status detected through the detection period

    @property
    def doc(self):
        return self._doc

    @property
    def doc_max(self):
        return self._doc_max

    @property
    def doc_min(self):
        return self._doc_min

    @property
    def status(self):
        return self._status

    @property
    def count_open(self):
        return self._count_open

    @property
    def count_closed(self):
        return self._count_closed

    @status.setter
    def status(self, value):
        self._status = value

    @doc_max.setter
    def doc_max(self, value):
        self._doc_max = value

    @doc_min.setter
    def doc_min(self, value):
        self._doc_min = value

    @doc.setter
    def doc(self, value):
        self._doc = value

    @count_open.setter
    def count_open(self, value):
        self._count_open = value

    @count_closed.setter
    def count_closed(self, value):
        self._count_closed = value


class MyGUI:

    def __init__(self):

        # Store all Window objects in a dictionary
        self.windows = {}
        for i in range(1, 28):
            window_no = f"upper_{i}"
            self.windows[window_no] = Window(doc=0, doc_min=0, doc_max=100, status="not_detected", count_open=0,
                                             count_closed=0)

        for i in range(1, 28):
            window_no = f"middle_{i}"
            self.windows[window_no] = Window(doc=0, doc_min=0, doc_max=100, status="not_detected", count_open=0,
                                             count_closed=0)

        for i in range(1, 19):
            window_no = f"lower_{i}"
            self.windows[window_no] = Window(doc=0, doc_min=0, doc_max=100, status="not_detected", count_open=0,
                                             count_closed=0)

        self.ui = QUiLoader().load('Detection Platform.ui')
        super().__init__()
        self.ui.comboBox.addItem("open")
        self.ui.comboBox.addItem("closed")
        self.ui.comboBox.addItem("blocked")
        self.ui.comboBox.addItem("not_detected")
        self.ui.comboBox.setCurrentIndex(3)
        self.ui.uploadButton.clicked.connect(self.on_uploadButton_clicked)
        self.ui.detectButton.clicked.connect(self.on_detectButton_clicked)
        self.ui.localizeButton.clicked.connect(self.on_localizeButton_clicked)
        self.ui.saveButton.clicked.connect(self.on_saveButton_clicked)
        self.ui.exportButton.clicked.connect(self.on_exportButton_clicked)

        # define and link button instances
        for i in range(1, 28):
            button_name = f"upper_{i}_Button"
            window_no = f"upper_{i}"
            button = getattr(self.ui, button_name)
            button.clicked.connect(functools.partial(self.on_window_clicked, window_no, button_name))

        for i in range(1, 28):
            button_name = f"middle_{i}_Button"
            window_no = f"middle_{i}"
            button = getattr(self.ui, button_name)
            button.clicked.connect(functools.partial(self.on_window_clicked, window_no, button_name))

        for i in range(1, 19):
            button_name = f"lower_{i}_Button"
            window_no = f"lower_{i}"
            button = getattr(self.ui, button_name)
            button.clicked.connect(functools.partial(self.on_window_clicked, window_no, button_name))

        self.output_size = 640
        self.img2predict = ""
        self.device = 'cpu'
        self.model = self.model_load(weights="weights/best.pt", device=self.device)

    def model_load(self, weights="",  # model.pt path(s)
                   device='',  # cuda device, i.e. 0 or 0,1,2,3 or cpu
                   half=False,  # use FP16 half-precision inference
                   dnn=False,  # use OpenCV DNN for ONNX inference
                   ):
        device = select_device(device)
        half &= device.type != 'cpu'  # half precision only supported on CUDA
        device = select_device(device)
        model = DetectMultiBackend(weights, device=device, dnn=dnn)
        stride, names, pt, jit, onnx = model.stride, model.names, model.pt, model.jit, model.onnx
        # Half
        half &= pt and device.type != 'cpu'  # half precision only supported by PyTorch on CUDA
        if pt:
            model.model.half() if half else model.model.float()
        print("Detection model deployed!")
        return model

    def update_graph(self):
        # Draw a pie chart
        status_count = {"open": 0, "closed": 0, "blocked": 0, "not_detected": 0}
        for window in self.windows.values():
            status_count[window.status] += 1

        # Set the sizes to the count of each status
        sizes = [status_count["open"], status_count["closed"], status_count["blocked"], status_count["not_detected"]]
        labels = ["Open", "Closed", "Blocked", "Not Detected"]
        colors = ["#87CEEB", "#FFA500", "#FF4500", "white"]
        explode = [0.02] * len(sizes)

        fig, ax = plt.subplots()
        wedges, _, autotexts = ax.pie(sizes, labels=labels, colors=colors, explode=explode, startangle=90,
                                      wedgeprops={"width": 0.4, "edgecolor": "black", "linewidth": 2},
                                      textprops={"fontsize": 14, "fontweight": "bold"},
                                      autopct="%1.1f%%")

        # Set font size and style for percentage labels
        for autotext in autotexts:
            autotext.set_fontsize(20)
            autotext.set_bbox(
                dict(facecolor='white', edgecolor='white', alpha=0.7))  # Set background color and transparency

        ax.axis("equal")  # Make the pie chart a perfect circle
        ax.set_title("Window Status", fontweight="bold", fontsize=20, loc="left")
        plt.tight_layout()

        # Save the pie chart as a pixmap
        buffer = io.BytesIO()
        plt.savefig(buffer, format="png")
        buffer.seek(0)
        pixmap = QPixmap()
        pixmap.loadFromData(buffer.getvalue())
        plt.savefig("images/result/pie_chart.png")

        self.ui.label_pie.setPixmap(pixmap)
        self.ui.label_pie.setScaledContents(True)

        # Select all the Window instances whose status is "closed" and get their doc attribute
        docs = [window.doc for window in self.windows.values() if window.status == "closed"]

        # Count the number of times each doc value appears
        bins = np.arange(0, 110, 10)
        hist, _ = np.histogram(docs, bins=bins)

        # Draw a bar chart
        fig, ax = plt.subplots()
        ax.bar(bins[:-1], hist, width=8, align="edge", color="#87CEEB", edgecolor="black", linewidth=2)
        ax.set_xticks(bins)
        ax.set_xlabel("Doc")
        ax.set_ylabel("Count")
        ax.set_ylim(bottom=0)
        ax.set_title("DoC Distribution", fontweight="bold", fontsize=20, loc="left")

        # Add count labels on each bar
        for i, v in enumerate(hist):
            if v != 0:
                ax.bar(bins[i], v, width=8, align="edge", color="#87CEEB", edgecolor="black", linewidth=2)
                ax.text(bins[i] + 4, v - 0.5, str(v), fontweight="bold", ha='center', va='bottom',
                        bbox=dict(facecolor='white', edgecolor='white', alpha=0.7))

        # Save the figure as a pixmap
        buffer = io.BytesIO()
        plt.savefig(buffer, format="png")
        buffer.seek(0)
        pixmap = QPixmap()
        pixmap.loadFromData(buffer.getvalue())
        plt.savefig("images/result/bar_chart.png")

        # Display the pixmap on the label_line
        self.ui.label_line.setPixmap(pixmap)
        self.ui.label_line.setScaledContents(True)

    def on_uploadButton_clicked(self):

        # Initialize charts and text
        empty_image = QPixmap(QSize(0, 0))
        empty_image.fill(Qt.transparent)
        self.ui.result_img_1.setPixmap(empty_image)
        self.ui.result_img_2.setPixmap(empty_image)
        self.ui.result_img_3.setPixmap(empty_image)
        self.ui.label_line.setPixmap(empty_image)
        self.ui.label_pie.setPixmap(empty_image)
        self.ui.windowText.setText("")
        self.ui.comboBox.setCurrentIndex(3)

        # Add window instances and initialize corresponding buttons
        for i in range(1, 28):
            button_name = f"upper_{i}_Button"
            window_no = f"upper_{i}"
            button = getattr(self.ui, button_name)
            button.setStyleSheet("background-color: #FFFFFF;")
            self.windows[window_no] = Window(doc=0, doc_min=0, doc_max=100, status="not_detected", count_open=0,
                                             count_closed=0)

        for i in range(1, 28):
            button_name = f"middle_{i}_Button"
            window_no = f"middle_{i}"
            button = getattr(self.ui, button_name)
            button.setStyleSheet("background-color: #FFFFFF;")
            self.windows[window_no] = Window(doc=0, doc_min=0, doc_max=100, status="not_detected", count_open=0,
                                             count_closed=0)

        for i in range(1, 19):
            button_name = f"lower_{i}_Button"
            window_no = f"lower_{i}"
            button = getattr(self.ui, button_name)
            button.setStyleSheet("background-color: #FFFFFF;")
            self.windows[window_no] = Window(doc=0, doc_min=0, doc_max=100, status="not_detected", count_open=0,
                                             count_closed=0)

        # read images
        fileNames, _ = QFileDialog.getOpenFileNames(None, 'Choose files', '', '*.jpg *.png *.tif *.jpeg')
        if fileNames:

            # If more than 3 images are uploaded, display the warning
            if len(fileNames) > 3:
                QMessageBox.warning(None, 'Warning', 'Maximum three images can be uploaded at once!')
                return

            self.img2predict = []
            self.window_no = []
            last_upload_time = None

            # clear previous tmp images
            tmp_folder = "images/tmp"
            if os.path.exists(tmp_folder):
                shutil.rmtree(tmp_folder)
            os.makedirs(tmp_folder)

            # Reset labels
            self.ui.label0.clear()
            self.ui.label1.clear()
            self.ui.label2.clear()

            for i, fileName in enumerate(fileNames):
                suffix = fileName.split(".")[-1]
                save_path = osp.join("images/tmp", f"tmp_upload_{i}." + suffix)
                shutil.copy(fileName, save_path)
                im = cv2.imread(save_path)
                # Read the labeled txt file, note that it should have the same name as
                # the corresponding image and be placed in the same folder.
                txt_filename, ext = os.path.splitext(fileName)
                txt_path = f"{txt_filename}.txt"

                # If no text file found, display the warning.
                if not os.path.exists(txt_path):
                    QMessageBox.warning(None, 'Warning', 'No txt file found!')
                    return

                # Obtain photo taken time, usually the file name contains the time.
                # If not, will use the recorded taken time in the file properties.
                # This will also check if the images are taken in one hour to ensure the detection is valid.
                basename = os.path.basename(fileName)
                try:
                    nameWithoutExt = os.path.splitext(basename)[0]
                    printable = set(string.printable)
                    nameWithoutExt = ''.join(filter(lambda x: x in printable, nameWithoutExt))
                    date_time = datetime.strptime(nameWithoutExt, '%Y %m %d %H;%M;%S')
                    print(date_time)
                    if last_upload_time is None:
                        last_upload_time = date_time
                    if last_upload_time is not None and abs(date_time - last_upload_time) > timedelta(
                            hours=1):
                        QMessageBox.warning(None, 'Warning',
                                            'Make sure uploading images during one same period!')
                        return
                    last_upload_time = date_time
                    self.ui.dateTimeEdit.setDateTime(date_time)

                # If the file name is not standard (contains photo taken time)
                except ValueError:

                    exif_data = Image.open(save_path)._getexif()
                    # initialize the time displayed
                    date_time = QDateTime(2000, 1, 1, 0, 0, 0)

                    if exif_data:
                        date_time = exif_data.get(36867)
                        if date_time:
                            date_time = datetime.strptime(date_time, '%Y:%m:%d %H:%M:%S')
                            if last_upload_time is None:
                                last_upload_time = date_time
                            if last_upload_time is not None and abs(date_time - last_upload_time) > timedelta(
                                    hours=1):
                                QMessageBox.warning(None, 'Warning',
                                                    'Make sure uploading images during one same period!')
                                return
                            last_upload_time = date_time

                    self.ui.dateTimeEdit.setDateTime(date_time)

                with open(txt_path, 'r') as file:
                    index = 0
                    for line in file:
                        parts = line.strip().split(' ')
                        self.window_no.append(str(parts[0]))
                        window_no = parts[0]
                        coord_upper_left = eval(parts[1])
                        coord2_upper_right = eval(parts[2])
                        coord3_lower_right = eval(parts[3])
                        coord4_lower_left = eval(parts[4])

                        # The coordinates of the four points in the source image:
                        # top-left, top-right, bottom-right, and bottom-left.
                        src_points = np.float32([np.float32(coord_upper_left), np.float32(coord2_upper_right),
                                                 np.float32(coord3_lower_right), np.float32(coord4_lower_left)])

                        # Coordinates of the four points in the destination image
                        dst_points = np.float32([[0, 0], [640, 0], [640, 640], [0, 640]])

                        # Calculate the perspective transformation matrix
                        M = cv2.getPerspectiveTransform(src_points, dst_points)

                        # Perform perspective transformation on the source image
                        dst_img = cv2.warpPerspective(im, M, (640, 640))

                        # Cropping images and add to img2predict for further detection
                        cv2.imwrite(f"images/tmp/warp_img_{i}_{index}_{window_no}.jpg", dst_img)
                        self.img2predict.append(f"images/tmp/warp_img_{i}_{index}_{window_no}.jpg")
                        index += 1

                # Display images in GUI
                im = cv2.cvtColor(im, cv2.COLOR_BGR2RGB)
                height, width, channels = im.shape
                bytesPerLine = channels * width
                qimage = QImage(im.data, width, height, bytesPerLine, QImage.Format_RGB888)
                pixmap = QPixmap.fromImage(qimage)
                if i == 0:
                    self.ui.label0.setPixmap(pixmap)
                    self.ui.label0.setScaledContents(True)
                elif i == 1:
                    self.ui.label1.setPixmap(pixmap)
                    self.ui.label1.setScaledContents(True)
                elif i == 2:
                    self.ui.label2.setPixmap(pixmap)
                    self.ui.label2.setScaledContents(True)

    def on_detectButton_clicked(self):
        model = self.model
        output_size = self.output_size
        imgsz = [640, 640]  # inference size (pixels)
        max_det = 10  # maximum detections per image
        device = self.device  # cuda device, i.e. 0 or 0,1,2,3 or cpu
        view_img = False  # show results
        save_txt = False  # save results to *.txt
        save_conf = False  # save confidences in --save-txt labels
        save_crop = False  # save cropped prediction boxes
        nosave = False  # do not save images/videos
        classes = None  # filter by class: --class 0, or --class 0 2 3
        agnostic_nms = False  # class-agnostic NMS
        augment = False  # ugmented inference
        visualize = False  # visualize features
        line_thickness = 3  # bounding box thickness (pixels)
        hide_labels = False  # hide labels
        hide_conf = False  # hide confidences
        half = False  # use FP16 half-precision inference
        dnn = False  # use OpenCV DNN for ONNX inference
        index = 0

        # Set confidence threshold, default = 0.25
        try:
            set_conf = float(self.ui.confidenceEdit.toPlainText())
            if 0 <= set_conf <= 1:
                conf_thres = set_conf
            else:
                QMessageBox.warning(None, "Invalid Threshold", "Confidence threshold must be a number between 0 and 1.")
                return
        except ValueError:
            if not self.ui.confidenceEdit.toPlainText():
                conf_thres = 0.25
            else:
                QMessageBox.warning(None, "Invalid Threshold", "Confidence threshold must be a number between 0 and 1.")
                return

        # Set NMS IOU threshold, default = 0.45
        try:
            set_IoU = float(self.ui.IoUEdit.toPlainText())
            if 0 <= set_IoU <= 1:
                iou_thres = set_IoU
            else:
                QMessageBox.warning(None, "Invalid Threshold", "IoU threshold must be a number between 0 and 1.")
                return
        except ValueError:
            if not self.ui.confidenceEdit.toPlainText():
                iou_thres = 0.45
            else:
                QMessageBox.warning(None, "Invalid Threshold", "IoU threshold must be a number between 0 and 1.")
                return

        self.ui.progressBar.reset()
        num_images = len(self.img2predict)

        # clear previous results
        result_folder = "images/result"
        if os.path.exists(result_folder):
            shutil.rmtree(result_folder)
        os.makedirs(result_folder)

        # Iterate all cropped images
        for source in self.img2predict:
            # Initialize parameters
            glass_detected = False
            shutter_detected = False
            glass_upper = 640
            shutter_lower = 0
            window_no = self.window_no[index]
            button = f"{window_no}_Button"

            # Read text file and link window instances with buttons
            try:
                button_instance = getattr(self.ui, button)
            except Exception:
                QMessageBox.warning(None, "Warning",
                                    "Please ensure your txt file is standard (window no. should be like 'upper_1')",
                                    QMessageBox.Ok)
                return

            if source == "":
                QMessageBox.warning(self, "warning", "Please upload images before detection")
            else:
                source = str(source)
                device = select_device(self.device)
                stride, names, pt, jit, onnx = model.stride, model.names, model.pt, model.jit, model.onnx
                imgsz = check_img_size(imgsz, s=stride)  # check image size
                save_img = not nosave and not source.endswith('.txt')  # save inference images

                dataset = LoadImages(source, img_size=imgsz, stride=stride, auto=pt and not jit)
                bs = 1  # batch_size
                vid_path, vid_writer = [None] * bs, [None] * bs
                # Run inference
                if pt and device.type != 'cpu':
                    model(torch.zeros(1, 3, *imgsz).to(device).type_as(next(model.model.parameters())))  # warmup
                dt, seen = [0.0, 0.0, 0.0], 0
                for path, im, im0s, vid_cap, s in dataset:
                    t1 = time_sync()
                    im = torch.from_numpy(im).to(device)
                    im = im.half() if half else im.float()  # uint8 to fp16/32
                    im /= 255  # 0 - 255 to 0.0 - 1.0
                    if len(im.shape) == 3:
                        im = im[None]  # expand for batch dim
                    t2 = time_sync()
                    dt[0] += t2 - t1
                    # Inference
                    # visualize = increment_path(save_dir / Path(path).stem, mkdir=True) if visualize else False
                    pred = model(im, augment=augment, visualize=visualize)
                    t3 = time_sync()
                    dt[1] += t3 - t2
                    # NMS
                    pred = non_max_suppression(pred, conf_thres, iou_thres, classes, agnostic_nms, max_det=max_det)
                    dt[2] += time_sync() - t3
                    # Second-stage classifier (optional)
                    # pred = utils.general.apply_classifier(pred, classifier_model, im, im0s)
                    # Process predictions
                    for i, det in enumerate(pred):  # per image
                        seen += 1
                        p, im0, frame = path, im0s.copy(), getattr(dataset, 'frame', 0)
                        p = Path(p)  # to Path
                        s += '%gx%g ' % im.shape[2:]  # print string
                        gn = torch.tensor(im0.shape)[[1, 0, 1, 0]]  # normalization gain whwh
                        imc = im0.copy() if save_crop else im0  # for save_crop
                        annotator = Annotator(im0, line_width=line_thickness, example=str(names))
                        img_height = im0.shape[0]
                        if len(det):

                            # Rescale boxes from img_size to im0 size
                            det[:, :4] = scale_coords(im.shape[2:], det[:, :4], im0.shape).round()

                            # Print results
                            for c in det[:, -1].unique():
                                n = (det[:, -1] == c).sum()  # detections per class
                                s += f"{n} {names[int(c)]}{'s' * (n > 1)}, "  # add to string

                            # Write results
                            for *xyxy, conf, cls in reversed(det):

                                # Glass detected
                                if cls == 0:

                                    glass_detected = True
                                    glass_upper = min(glass_upper, int(xyxy[1]))

                                    if save_img or save_crop or view_img:  # Add bbox to image
                                        c = int(cls)  # integer class
                                        label = None if hide_labels else (
                                            names[c] if hide_conf else f'{names[c]} {conf:.2f}')
                                        annotator.box_label(xyxy, label, color=colors(c, True))

                                # Roller shutter detected
                                if cls == 1:

                                    shutter_detected = True
                                    shutter_lower = max(shutter_lower, int(xyxy[3]))

                                    if save_img or save_crop or view_img:  # Add bbox to image
                                        c = int(cls)  # integer class
                                        label = None if hide_labels else (
                                            names[c] if hide_conf else f'{names[c]} {conf:.2f}')
                                        annotator.box_label(xyxy, label, color=colors(c, True))

                            if glass_detected and not shutter_detected:

                                # preset threshold 100 for upper window frame
                                if glass_upper < 100:
                                    self.windows[window_no].count_open += 1
                                    if self.windows[window_no].count_open > self.windows[window_no].count_closed:
                                        button_instance.setStyleSheet("background-color: #87CEEB;")
                                        self.windows[window_no].status = "open"
                                        print("-----------------------------------------------------------")
                                        print(f"Window No.{window_no}:")
                                        print(f"the shutter is open.")
                                        print("-----------------------------------------------------------")

                                else:
                                    if self.windows[window_no].status == "blocked" or self.windows[
                                        window_no].status == "not_detected":
                                        button_instance.setStyleSheet("background-color: #FF4500;")
                                        self.windows[window_no].status = "blocked"
                                        self.windows[window_no].doc_max = min(self.windows[window_no].doc_max,
                                                                              glass_upper * 100 / img_height)
                                        print("-----------------------------------------------------------")
                                        print(f"Window No.{window_no}:")
                                        print(f"the shutter is upper blocked.")
                                        print("-----------------------------------------------------------")

                            if not glass_detected and shutter_detected:

                                glass_upper = img_height
                                # preset threshold 500 for lower window frame
                                if shutter_lower < 500 and (self.windows[window_no].status == "blocked" or self.windows[
                                    window_no].status == "not_detected"):
                                    button_instance.setStyleSheet("background-color: #FF4500;")
                                    self.windows[window_no].status = "blocked"
                                    self.windows[window_no].doc_min = max(self.windows[window_no].doc_min,
                                                                          shutter_lower * 100 / img_height)

                                    print("-----------------------------------------------------------")
                                    print(f"Window No.{window_no}:")
                                    print(f"the shutter is lower blocked.")
                                    print("-----------------------------------------------------------")

                                else:
                                    self.windows[window_no].count_closed += 1
                                    if not self.windows[window_no].count_open > self.windows[window_no].count_closed:
                                        self.windows[window_no].status = "closed"
                                        if self.windows[window_no].doc == 0:
                                            self.windows[window_no].doc = max(self.windows[window_no].doc,
                                                                              shutter_lower * 100 / img_height)
                                        else:
                                            self.windows[window_no].doc = (self.windows[
                                                                               window_no].doc + shutter_lower * 100 / img_height) / 2

                                        button_instance.setStyleSheet("background-color: #FFA500;")
                                        print("-----------------------------------------------------------")
                                        print(f"Window No.{window_no}:")
                                        print(f"Closing degree of the shutter: {self.windows[window_no].doc:.2f}% closed")
                                        print("-----------------------------------------------------------")

                            if glass_detected and shutter_detected:

                                self.windows[window_no].count_closed += 1
                                if not self.windows[window_no].count_open > self.windows[window_no].count_closed:
                                    self.windows[window_no].status = "closed"
                                    if self.windows[window_no].doc == 0:
                                        self.windows[window_no].doc = max(self.windows[window_no].doc,
                                                                          shutter_lower * 100 / img_height)
                                    else:
                                        self.windows[window_no].doc = (self.windows[
                                                                           window_no].doc + shutter_lower * 100 / img_height) / 2

                                    button_instance.setStyleSheet("background-color: #FFA500;")

                                    print("-----------------------------------------------------------")
                                    print(f"Window No.{window_no}:")
                                    print(f"Closing degree of the shutter: {self.windows[window_no].doc:.2f}% closed")
                                    print("-----------------------------------------------------------")

                            else:
                                if self.windows[window_no].status == "blocked" or self.windows[
                                    window_no].status == "not_detected":
                                    button_instance.setStyleSheet("background-color: #FF4500;")
                                    self.windows[window_no].status = "blocked"
                                    self.windows[window_no].doc_min = max(self.windows[window_no].doc_min,
                                                                          shutter_lower * 100 / img_height)
                                    self.windows[window_no].doc_max = min(self.windows[window_no].doc_max,
                                                                          glass_upper * 100 / img_height)

                                    print("-----------------------------------------------------------")
                                    print(f"Window No.{window_no}:")
                                    print(f"the window is middle blocked.")
                                    print("-----------------------------------------------------------")
                        else:
                            if self.windows[window_no].status == "blocked" or self.windows[
                                window_no].status == "not_detected":
                                button_instance.setStyleSheet("background-color: #FF4500;")
                                self.windows[window_no].status = "blocked"

                                print("-----------------------------------------------------------")
                                print(f"Window No.{window_no}:")
                                print(f"the shutter is totally blocked.")
                                print("-----------------------------------------------------------")

                        LOGGER.info(f'{s}Done. ({t3 - t2:.3f}s)')
                        # Stream results
                        im0 = annotator.result()
                        resize_scale = output_size / im0.shape[0]
                        im0 = cv2.resize(im0, (0, 0), fx=resize_scale, fy=resize_scale)
                        file_name, file_ext = os.path.splitext(os.path.basename(p))
                        result_file_name = window_no + "_result" + file_ext
                        i = 1
                        # Save result images
                        while os.path.exists(os.path.join("images/result", result_file_name)):
                            result_file_name = window_no + f"_result_{i}" + file_ext
                            i += 1
                        cv2.imwrite(os.path.join("images/result", result_file_name), im0)
                        progress = (index + 1) * 100 // num_images
                        self.ui.progressBar.setValue(progress)
                        index += 1

        self.update_graph()

    def on_exportButton_clicked(self):
        # create a new Excel file
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # add information
        sheet['A1'] = 'window_no'
        sheet['B1'] = 'status'
        sheet['C1'] = 'degree of closure(DoC)'
        sheet['D1'] = 'max DoC (if blocked)'
        sheet['E1'] = 'min DoC (if blocked)'

        for i in range(1, 28):
            row_num = i + 1
            sheet['A{}'.format(row_num)] = 'upper_{}'.format(i)
            sheet['B{}'.format(row_num)] = self.windows['upper_{}'.format(i)].status
            sheet['C{}'.format(row_num)] = 0 if self.windows['upper_{}'.format(i)].status == "open" else \
                float(self.windows['upper_{}'.format(i)].doc) / 100 if \
                    self.windows['upper_{}'.format(i)].status == "closed" else ''
            sheet['D{}'.format(row_num)] = float(self.windows['upper_{}'.format(i)].doc_max) / 100 if \
                self.windows['upper_{}'.format(i)].status == "blocked" else ''
            sheet['E{}'.format(row_num)] = float(self.windows['upper_{}'.format(i)].doc_min) / 100 if \
                self.windows['upper_{}'.format(i)].status == "blocked" else ''

        for i in range(1, 28):
            row_num = i + 28 + 1
            sheet['A{}'.format(row_num)] = 'middle_{}'.format(i)
            sheet['B{}'.format(row_num)] = self.windows['middle_{}'.format(i)].status
            sheet['C{}'.format(row_num)] = 0 if self.windows['middle_{}'.format(i)].status == "open" else \
                float(self.windows['middle_{}'.format(i)].doc) / 100 if \
                    self.windows['middle_{}'.format(i)].status == "closed" else ''
            sheet['D{}'.format(row_num)] = float(self.windows['middle_{}'.format(i)].doc_max) / 100 if \
                self.windows['middle_{}'.format(i)].status == "blocked" else ''
            sheet['E{}'.format(row_num)] = float(self.windows['middle_{}'.format(i)].doc_min) / 100 if \
                self.windows['middle_{}'.format(i)].status == "blocked" else ''

        for i in range(1, 18):
            row_num = i + 56 + 1
            sheet['A{}'.format(row_num)] = 'lower_{}'.format(i)
            sheet['B{}'.format(row_num)] = self.windows['lower_{}'.format(i)].status
            sheet['C{}'.format(row_num)] = 0 if self.windows['lower_{}'.format(i)].status == "open" else \
                float(self.windows['lower_{}'.format(i)].doc) / 100 if \
                    self.windows['lower_{}'.format(i)].status == "closed" else ''
            sheet['D{}'.format(row_num)] = float(self.windows['lower_{}'.format(i)].doc_max) / 100 if \
                self.windows['lower_{}'.format(i)].status == "blocked" else ''
            sheet['E{}'.format(row_num)] = float(self.windows['lower_{}'.format(i)].doc_min) / 100 if \
                self.windows['lower_{}'.format(i)].status == "blocked" else ''

        for col in ['C', 'D', 'E']:
            for i in range(2, 84):
                cell = sheet['{}{}'.format(col, i)]
                if cell.value != '':
                    cell.number_format = '0.00%'

        # Modify table display style
        for col in sheet.iter_cols(min_row=1, max_row=1):
            for cell in col:
                cell.font = Font(name='timesnewroman', size=11, bold=True)
        sheet.row_dimensions[1].height = 25

        for column_index in range(1, 6):
            header_cell = sheet.cell(row=1, column=column_index)
            header_cell.font = Font(name='Times New Roman', bold=True, size=11)
            header_cell.alignment = Alignment(horizontal='center')
            header_cell.border = Border(bottom=Side(style='thin', color='000000'))

        for column in sheet.columns:
            column_length = max(len(str(cell.value)) for cell in column)
            column_letter = get_column_letter(column[0].column)
            sheet.column_dimensions[column_letter].width = column_length + 4

        color_map = {
            'open': '87CEEB',
            'closed': 'FFA500',
            'blocked': 'FF4500'
        }

        for i in range(1, 28):
            row_num = i + 1
            cell = sheet['B{}'.format(row_num)]
            status = self.windows['upper_{}'.format(i)].status
            if status in color_map:
                cell.fill = PatternFill(start_color=color_map[status], end_color=color_map[status], fill_type='solid')

        for i in range(1, 28):
            row_num = i + 29
            cell = sheet['B{}'.format(row_num)]
            status = self.windows['middle_{}'.format(i)].status
            if status in color_map:
                cell.fill = PatternFill(start_color=color_map[status], end_color=color_map[status], fill_type='solid')
        for i in range(1, 19):
            row_num = i + 57
            cell = sheet['B{}'.format(row_num)]
            status = self.windows['lower_{}'.format(i)].status
            if status in color_map:
                cell.fill = PatternFill(start_color=color_map[status], end_color=color_map[status], fill_type='solid')

        # Add charts to the table
        img1_pil = Image.open('images/result/bar_chart.png')
        img1 = XLImage(img1_pil)
        img1.width = 300
        img1.height = 220
        sheet.add_image(img1, 'G16')

        img2_pil = Image.open('images/result/pie_chart.png')
        img2 = XLImage(img2_pil)
        img2.width = 300
        img2.height = 220
        sheet.add_image(img2, 'G3')
        # Set default file name as the photo taken time
        title = self.ui.dateTimeEdit.text().replace('/', '.').replace(':', ';')
        root = tk.Tk()
        root.withdraw()

        file_path = filedialog.asksaveasfilename(defaultextension='.xlsx', initialfile='{}.xlsx'.format(title))

        if file_path:
            workbook.save(file_path)
            messagebox.showinfo('Export Successful', 'File has been exported successfully!')

    def on_window_clicked(self, window_no, button):
        # Initialize the display interface
        self.ui.windowText.clear()
        self.ui.result_img_1.clear()
        self.ui.result_img_2.clear()
        self.ui.result_img_3.clear()

        status = self.windows[window_no].status
        if status == "open":
            self.ui.comboBox.setCurrentIndex(0)
        elif status == "closed":
            self.ui.comboBox.setCurrentIndex(1)
        elif status == "blocked":
            self.ui.comboBox.setCurrentIndex(2)
        elif status == "not_detected":
            self.ui.comboBox.setCurrentIndex(3)

        folder_path = "images/result/"
        # Search for files starting with "{window_no}_result"
        target_prefix = f"{window_no}_result"
        i = 0
        for file_name in os.listdir(folder_path):
            if file_name.startswith(target_prefix):
                # Construct the complete path of the file
                file_path = os.path.join(folder_path, file_name)
                # Read the image and display it on the corresponding QLabel
                im = cv2.imread(file_path)
                im = cv2.cvtColor(im, cv2.COLOR_BGR2RGB)
                height, width, channels = im.shape
                bytesPerLine = channels * width
                qimage = QImage(im.data, width, height, bytesPerLine, QImage.Format_RGB888)
                pixmap = QPixmap.fromImage(qimage)
                if i == 0:
                    self.ui.result_img_1.setPixmap(pixmap)
                    self.ui.result_img_1.setScaledContents(True)
                elif i == 1:
                    self.ui.result_img_2.setPixmap(pixmap)
                    self.ui.result_img_2.setScaledContents(True)
                elif i == 2:
                    self.ui.result_img_3.setPixmap(pixmap)
                    self.ui.result_img_3.setScaledContents(True)
                # Only display the first 3 images
                if i >= 2:
                    break
                i += 1

        if self.windows[window_no].status == "closed":
            self.ui.windowText.setText(f"{window_no} is closed. ({self.windows[window_no].doc:.2f}%).")
        if self.windows[window_no].status == "open":
            self.ui.windowText.setText(f"{window_no} is open.")
        if self.windows[window_no].status == "blocked":
            self.ui.windowText.setText(
                f"{window_no} blocked. ({self.windows[window_no].doc_min:.2f}% to {self.windows[window_no].doc_max:.2f}%)")
        if self.windows[window_no].status == "not_detected":
            self.ui.windowText.setText(f"{window_no} has not been detected yet.")

    def on_localizeButton_clicked(self):

        text = self.ui.windowText.toPlainText()
        window_no = text.split()[0]
        current_tab = self.ui.tabWidget_2.currentIndex()

        if self.ui.comboBox.currentIndex() == 1:  # If chosen status is "closed"

            # Get the path of the currently selected image
            if current_tab == 0:
                pixmap = self.ui.result_img_1.pixmap()
            elif current_tab == 1:
                pixmap = self.ui.result_img_2.pixmap()
            else:
                pixmap = self.ui.result_img_3.pixmap()

            dialog = LocalizeDialogClosed(pixmap)
            if dialog.exec_() == QDialog.Accepted:
                shutter_pos = dialog.shutter_pos
                self.ui.windowText.setText(f"{window_no} is closed. ({shutter_pos / 6.4:.2f}%).")

        if self.ui.comboBox.currentIndex() == 2:  # If chosen status is "blocked"

            if current_tab == 0:
                pixmap = self.ui.result_img_1.pixmap()
            elif current_tab == 1:
                pixmap = self.ui.result_img_2.pixmap()
            else:
                pixmap = self.ui.result_img_3.pixmap()

            dialog = LocalizeDialogBlocked(pixmap)
            if dialog.exec_() == QDialog.Accepted:
                shutter_pos_max = dialog.shutter_pos_max
                shutter_pos_min = dialog.shutter_pos_min
                self.ui.windowText.setText(
                    f"{window_no} blocked. ({shutter_pos_min / 6.4:.2f}% to {shutter_pos_max / 6.4:.2f}%)")

        elif self.ui.comboBox.currentIndex() == 0 or self.ui.comboBox.currentIndex() == 3:
            msg_box = QMessageBox()
            msg_box.setText("You can only localize the shutter when you set its status as closed or blocked!")
            msg_box.setIcon(QMessageBox.Warning)
            msg_box.exec_()
            return

    def on_saveButton_clicked(self):

        text = self.ui.windowText.toPlainText()
        window_no = text.split()[0]
        button = f"{window_no}_Button"
        button_instance = getattr(self.ui, button)

        self.windows[window_no].status = self.ui.comboBox.currentText()

        if self.ui.comboBox.currentIndex() == 1:
            # Get doc from windowText displayed
            start = text.find("(") + 1
            end = text.find(")")
            try:
                new_doc = float(text[start:end].strip("%"))
            except Exception:
                QMessageBox.information(None, "Information",
                                        "DoC has been set to 100% (default), localizing is suggested before you save the changes!")
                new_doc = 100.0

            self.windows[window_no].doc = new_doc

        if self.ui.comboBox.currentIndex() == 2:
            # Get doc_min and doc_max from windowText displayed
            start = text.find("(") + 1
            end = text.find(")")
            try:
                new_doc_min = float(text[start:end].split()[0].strip("%"))
                new_doc_max = float(text[start:end].split()[2].strip("%"))
            except Exception:
                QMessageBox.information(None, "Information",
                                        "DoC has been set to (0-100%) (default), localizing is suggested before you save the changes!")
                new_doc_min = 0.0
                new_doc_max = 100.0
            self.windows[window_no].doc_min = new_doc_min
            self.windows[window_no].doc_max = new_doc_max

        # Modify the button appearance according to new status
        if self.windows[window_no].status == "open":
            self.ui.windowText.setText(f"{window_no} is open.")
            button_instance.setStyleSheet("background-color: #87CEEB;")
        if self.windows[window_no].status == "closed":
            self.ui.windowText.setText(f"{window_no} is closed. ({self.windows[window_no].doc:.2f}%).")
            button_instance.setStyleSheet("background-color: #FFA500;")
        if self.windows[window_no].status == "blocked":
            self.ui.windowText.setText(
                f"{window_no} blocked. ({self.windows[window_no].doc_min:.2f}% to {self.windows[window_no].doc_max:.2f}%)")
            button_instance.setStyleSheet("background-color: #FF4500;")
        if self.windows[window_no].status == "not_detected":
            button_instance.setStyleSheet("background-color: #FFFFFF;")
            self.ui.windowText.setText(f"{window_no} has not been detected yet.")

        self.update_graph()


class LocalizeDialogClosed(QDialog):
    def __init__(self, pixmap):
        super().__init__()
        self.setWindowTitle("Please localize the shutter bottom and press Enter to commit")
        self.setFixedSize(640, 640)

        self.label = QLabel(self)
        self.label.setGeometry(0, 0, self.width(), self.height())

        # Set the original image as the background of the QLabel
        self.label.setPixmap(pixmap)
        self.label.setScaledContents(True)

        # Register the mousePressEvent event handler function
        self.label.mousePressEvent = self.on_mouse_press_event

        # Register the keyPressEvent event handler function
        self.keyPressEvent = self.on_key_press_event

        # Create and configure a horizontal line
        self.h_line = QFrame(self)
        self.h_line.setFrameStyle(QFrame.HLine)
        self.h_line.setStyleSheet("color: green")
        self.h_line.setFixedWidth(self.width())
        self.h_line.hide()

    def on_mouse_press_event(self, event: QMouseEvent):  # Click to set the new boundary
        self.h_line.setStyleSheet("border: 3px solid green;")
        self.h_line.setFixedHeight(3)
        self.h_line.raise_()
        self.h_line.show()
        self.h_line.move(0, event.pos().y())
        self.shutter_pos = event.pos().y()

    def on_key_press_event(self, event):  # Press Enter to commit
        if event.key() == Qt.Key_Return or event.key() == Qt.Key_Enter:
            self.accept()


class LocalizeDialogBlocked(QDialog):
    def __init__(self, pixmap):
        super().__init__()
        self.setWindowTitle("Please click twice to localize the possible shutter top & bottom")
        self.setFixedSize(640, 640)
        self.click_count = 0
        self.shutter_pos_max = None
        self.shutter_pos_min = None

        self.label = QLabel(self)
        self.label.setGeometry(0, 0, self.width(), self.height())

        self.label.setPixmap(pixmap)
        self.label.setScaledContents(True)

        self.label.mousePressEvent = self.on_mouse_press_event

        self.h_line = QFrame(self)
        self.h_line.setFrameStyle(QFrame.HLine)
        self.h_line.setStyleSheet("color: green")
        self.h_line.setFixedWidth(self.width())
        self.h_line.hide()

    def on_mouse_press_event(self, event: QMouseEvent):  # Click twice to set the new possible area
        if self.click_count == 0:
            self.h_line.setStyleSheet("border: 3px solid green;")
            self.h_line.setFixedHeight(3)
            self.h_line.raise_()
            self.h_line.show()
            self.h_line.move(0, event.pos().y())
            self.shutter_pos_max = event.pos().y()
            self.shutter_pos_min = event.pos().y()
        else:
            pos_y = event.pos().y()
            self.shutter_pos_max = max(self.shutter_pos_max, pos_y)
            self.shutter_pos_min = min(self.shutter_pos_min, pos_y)

        self.click_count += 1

        if self.click_count == 2:
            # Close the window automatically after clicked twice
            self.accept()


app = QApplication([])
app.setWindowIcon(QIcon("images/logo.png"))
myGUI = MyGUI()
myGUI.ui.show()
app.exec_()
